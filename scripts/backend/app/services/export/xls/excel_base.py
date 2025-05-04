"""
@author maria
date: 2025-05-03
"""
from abc import ABC, abstractmethod
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fastapi.responses import StreamingResponse

# abstract base class (ABC)
class ExcelExporterBase(ABC):

    def __init__(self, titulo: str, subtitulo: str, nome_arquivo: str, filtros: Optional[dict] = None):
        self.titulo = titulo
        self.subtitulo = subtitulo
        self.nome_arquivo = nome_arquivo
        self.filtros = filtros or {}

    @abstractmethod
    def obter_dados(self) -> list[dict]:
        pass

    @abstractmethod
    def configurar_colunas_relatorio(self, ws):
        pass

    @abstractmethod
    def customiza_relatorio(self, ws, dados):
        pass

    @abstractmethod
    def configurar_colunas_filtros(self, ws_filtros):
        pass

    def gerar_excel(self):
        wb = Workbook()

        ws_filtros = wb.active
        ws_filtros.title = "Filtros"
        ws_filtros.append(["Filtro", "Valor"])
        ws_filtros["A1"].font = ws_filtros["B1"].font = Font(bold=True)

        for chave, valor in self.filtros.items():
            display_valor = valor if valor is not None else "-"
            ws_filtros.append([chave, display_valor])

        self.configurar_colunas_filtros(ws_filtros)

        ws = wb.create_sheet(title="Relatório")
        ws.merge_cells("A1:B1")
        ws["A1"] = self.titulo
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:B2")
        ws["A2"] = self.subtitulo
        ws["A2"].font = Font(size=12)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.append([])

        self.configurar_colunas_relatorio(ws)

        dados = self.obter_dados()
        for row in dados:
            ws.append(list(row.values()))

        wb._sheets = [ws_filtros] + [sheet for sheet in wb.worksheets if sheet != ws_filtros]

        self.customiza_relatorio(ws, dados)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{self.nome_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers=headers)


