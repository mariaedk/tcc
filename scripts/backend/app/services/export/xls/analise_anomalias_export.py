"""
@author maria
date: 2025-05-03
"""
from analise.models.resultado_analise import ResultadoAnaliseSchema
from app.models.enums import TipoMedicao
from .excel_base import ExcelExporterBase
from openpyxl.styles import Font

class AnomaliasExport(ExcelExporterBase):

    def __init__(self, resultado: ResultadoAnaliseSchema, tipo_medicao: TipoMedicao, filtros: dict = None):
        super().__init__(
            titulo="Vazão de Água ETA 2 com Destaques de Anomalias (m³/h)",
            subtitulo="Pontos em vermelho indicam comportamentos fora do padrão, identificados com IA.",
            nome_arquivo="analise_anomalias",
            filtros=filtros
        )
        self.resultado = resultado
        self.tipo_medicao = tipo_medicao

    def obter_dados(self):
        return [
            {
                "Data": ponto.data.strftime('%d/%m/%Y %H:%M') if self.tipo_medicao == TipoMedicao.HORA else ponto.data.strftime('%d/%m/%Y'),
                "Valor (m³/h)": ponto.valor,
                "Anomalia": "Sim" if ponto.is_anomalia else "Não"
            }
            for ponto in self.resultado.dados
        ]

    def configurar_colunas_relatorio(self, ws):
        ws.append(["Data", "Valor (m³/h)", "Anomalia"])
        ws['A4'].font = ws['B4'].font = ws['C4'].font = Font(bold=True)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 30

    def configurar_colunas_filtros(self, ws_filtros):
        ws_filtros.column_dimensions["A"].width = 20
        ws_filtros.column_dimensions["B"].width = 25

    def customiza_relatorio(self, ws, dados):
        total_anomalias = sum(1 for d in dados if d["Anomalia"] == "Sim")
        ws.append([])
        ws.append(["Total de Anomalias", total_anomalias])